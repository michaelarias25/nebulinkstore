import mysql.connector
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def exportar_postulaciones():
    # Conexión a MySQL
    conexion = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="nebulink_store"
    )

    cursor = conexion.cursor()
    cursor.execute("""
        SELECT 
            id,
            nombre,
            correo,
            telefono,
            estado,
            fecha_postulacion
        FROM postulaciones
        ORDER BY fecha_postulacion DESC
    """)

    filas = cursor.fetchall()

    # Crear carpeta de exportaciones
    base_dir = os.path.dirname(os.path.abspath(__file__))
    carpeta_exports = os.path.join(base_dir, "..", "exports")
    os.makedirs(carpeta_exports, exist_ok=True)

    ruta_archivo = os.path.join(carpeta_exports, "postulaciones.xlsx")

    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Postulaciones"

    # Encabezados con la nueva columna 'estado'
    encabezados = [
        "ID",
        "Nombre",
        "Correo",
        "Teléfono",
        "Estado",
        "Fecha Postulación"
    ]

    ws.append(encabezados)

    # Agregar filas
    for fila in filas:
        ws.append(list(fila))

    # Ajustar ancho de columnas
    for col in range(1, len(encabezados) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 25

    wb.save(ruta_archivo)

    cursor.close()
    conexion.close()

    print(f"Archivo Excel generado en: {ruta_archivo}")

if __name__ == "__main__":
    exportar_postulaciones()
